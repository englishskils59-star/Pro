# storage_manager.py
# WDI Visit Analytics Engine
# Persistent storage manager — saves/loads data from shared folder
# Supports multi-user access via network shared drive

import os
import json
import time
import shutil
import hashlib
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from contextlib import contextmanager

from utils import normalize_arabic, safe_str

# ═══════════════════════════════════════════════════════════════════
# CONFIG FILE — stores the shared folder path
# ═══════════════════════════════════════════════════════════════════

CONFIG_FILE = Path(__file__).parent / "wdi_config.json"
DEFAULT_DATA_DIR = Path(__file__).parent / "data"


def load_config() -> dict:
    """Load app configuration from local config file."""
    default = {
        "data_dir": str(DEFAULT_DATA_DIR),
        "app_version": "1.0",
        "last_updated": "",
    }
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                cfg = json.load(f)
                # Merge with defaults for any missing keys
                for k, v in default.items():
                    if k not in cfg:
                        cfg[k] = v
                return cfg
        except Exception:
            return default
    return default


def save_config(config: dict):
    """Save configuration to local config file."""
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        return False


def get_data_dir() -> Path:
    """Get the configured data directory path."""
    cfg = load_config()
    return Path(cfg.get("data_dir", str(DEFAULT_DATA_DIR)))


def set_data_dir(path: str) -> tuple[bool, str]:
    """
    Set the shared data directory path.
    Creates the directory if it doesn't exist.
    Returns (success, message).
    """
    try:
        p = Path(path.strip())
        p.mkdir(parents=True, exist_ok=True)

        # Test write access
        test_file = p / ".wdi_test"
        test_file.write_text("test")
        test_file.unlink()

        cfg = load_config()
        cfg["data_dir"] = str(p)
        save_config(cfg)
        return True, f"✅ تم تعيين مسار البيانات: {p}"
    except PermissionError:
        return False, f"❌ لا توجد صلاحية كتابة في: {path}"
    except Exception as e:
        return False, f"❌ خطأ: {e}"


# ═══════════════════════════════════════════════════════════════════
# FILE PATHS
# ═══════════════════════════════════════════════════════════════════

def _paths() -> dict:
    """Return all data file paths based on current data_dir."""
    d = get_data_dir()
    return {
        "metadata":     d / "metadata.json",
        "classified":   d / "classified_data.parquet",
        "journey":      d / "journey_data.parquet",
        "rep_kpi":      d / "rep_kpi_data.parquet",
        "overrides":    d / "overrides.parquet",
        "raw_backup":   d / "last_upload.xlsx",
        "custom_rules": d / "custom_rules.json",
        "name_merges":  d / "name_merges.parquet",
    }


# ═══════════════════════════════════════════════════════════════════
# WRITE LOCK — prevents two users on the shared folder from saving
# at the same moment and corrupting the parquet files
# ═══════════════════════════════════════════════════════════════════

@contextmanager
def _write_lock(timeout: float = 10.0):
    lock = get_data_dir() / ".wdi_lock"
    lock.parent.mkdir(parents=True, exist_ok=True)
    start = time.time()
    while True:
        try:
            fd = os.open(str(lock), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.write(fd, str(os.getpid()).encode())
            os.close(fd)
            break
        except FileExistsError:
            # A lock older than 60s is considered stale (crashed process)
            try:
                if time.time() - lock.stat().st_mtime > 60:
                    lock.unlink(missing_ok=True)
                    continue
            except OSError:
                pass
            if time.time() - start > timeout:
                raise TimeoutError("مستخدم آخر يقوم بالحفظ الآن — حاول بعد لحظات")
            time.sleep(0.3)
    try:
        yield
    finally:
        try:
            lock.unlink(missing_ok=True)
        except OSError:
            pass


# ═══════════════════════════════════════════════════════════════════
# METADATA
# ═══════════════════════════════════════════════════════════════════

def _load_metadata() -> dict:
    paths = _paths()
    if paths["metadata"].exists():
        try:
            with open(paths["metadata"], "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def _save_metadata(meta: dict):
    paths = _paths()
    paths["metadata"].parent.mkdir(parents=True, exist_ok=True)
    meta["last_saved"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    tmp = paths["metadata"].with_name("metadata.json.tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    os.replace(tmp, paths["metadata"])


# ═══════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════

def save_session(
    classified_df: pd.DataFrame,
    journey_df: pd.DataFrame,
    rep_kpi_df: pd.DataFrame,
    file_name: str = "",
    uploaded_file=None,
) -> tuple[bool, str]:
    """
    Save current session data to the shared data directory.
    Returns (success, message).
    """
    try:
        paths = _paths()
        data_dir = get_data_dir()
        data_dir.mkdir(parents=True, exist_ok=True)

        with _write_lock():
            # ── Save classified data ──
            _df_to_parquet(classified_df, paths["classified"])

            # ── Save journey data (drop internal columns) ──
            journey_save = journey_df.copy()
            if "_journey" in journey_save.columns:
                journey_save = journey_save.drop(columns=["_journey"])
            _df_to_parquet(journey_save, paths["journey"])

            # ── Save rep KPI ──
            if rep_kpi_df is not None and not rep_kpi_df.empty:
                _df_to_parquet(rep_kpi_df, paths["rep_kpi"])

            # ── Save raw file backup ──
            if uploaded_file is not None:
                try:
                    uploaded_file.seek(0)
                    with open(paths["raw_backup"], "wb") as f:
                        f.write(uploaded_file.read())
                except Exception:
                    pass

            # ── Save metadata (keep existing override count) ──
            meta = _load_metadata()
            meta.update({
                "file_name":       file_name,
                "total_records":   len(classified_df),
                "unique_customers":classified_df["Customer Name"].nunique() if "Customer Name" in classified_df.columns else 0,
                "unique_reps":     classified_df["Sales Rep Name"].nunique() if "Sales Rep Name" in classified_df.columns else 0,
                "date_range_start":str(classified_df["Visit Date"].min())[:10] if "Visit Date" in classified_df.columns else "",
                "date_range_end":  str(classified_df["Visit Date"].max())[:10] if "Visit Date" in classified_df.columns else "",
            })
            meta.setdefault("override_count", 0)
            _save_metadata(meta)

        return True, "✅ تم حفظ البيانات بنجاح"
    except Exception as e:
        return False, f"❌ خطأ في الحفظ: {e}"


# ═══════════════════════════════════════════════════════════════════
# LOAD
# ═══════════════════════════════════════════════════════════════════

def load_session() -> tuple[bool, dict]:
    """
    Load saved session from shared data directory.
    Returns (success, data_dict).
    data_dict keys: classified_df, journey_df, rep_kpi_df, metadata
    """
    try:
        paths = _paths()

        if not paths["classified"].exists():
            return False, {}

        classified_df = _parquet_to_df(paths["classified"])
        rep_kpi_df    = _parquet_to_df(paths["rep_kpi"]) if paths["rep_kpi"].exists() else pd.DataFrame()
        metadata      = _load_metadata()

        # Unify rep-name spelling on data saved before this feature existed
        if "Sales Rep Name" in classified_df.columns:
            from utils import clean_rep_name
            classified_df["Sales Rep Name"] = classified_df["Sales Rep Name"].apply(clean_rep_name)

        # Apply saved name merges, then saved overrides
        classified_df = apply_name_merges(classified_df)
        _rekey_overrides_if_needed(classified_df)
        classified_df, override_count = _apply_saved_overrides(classified_df)

        # Rebuild the journey instead of loading the stale saved copy:
        # "Days Since Last Visit" was frozen at save time, and overrides/merges
        # may have changed customer statuses since.
        from classification_engine import build_customer_journey
        journey_df = build_customer_journey(classified_df)

        return True, {
            "classified_df": classified_df,
            "journey_df":    journey_df,
            "rep_kpi_df":    rep_kpi_df,
            "metadata":      metadata,
            "override_count":override_count,
        }
    except Exception as e:
        return False, {"error": str(e)}


def apply_saved_overrides(classified_df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Public wrapper — used by the upload pipeline so manual classifications
    survive a new file upload without waiting for an app restart."""
    return _apply_saved_overrides(classified_df)


def has_saved_data() -> bool:
    """Check if saved data exists in the data directory."""
    return _paths()["classified"].exists()


def get_saved_metadata() -> dict:
    """Get metadata of saved session without loading full data."""
    return _load_metadata()


# ═══════════════════════════════════════════════════════════════════
# OVERRIDE SYSTEM
# ═══════════════════════════════════════════════════════════════════

VALID_STATUSES = [
    "Current Customer",
    "Potential Customer",
    "Target Customer",
    "New Customer",
    "Former Customer",
    "Not Interested",
    "No Meeting",
]


# ═══════════════════════════════════════════════════════════════════
# STABLE OVERRIDE KEY
# Overrides used to be stored by row number, which silently applied
# old classifications to the wrong rows after a new file upload.
# The key below identifies the VISIT itself, so overrides survive
# re-uploads correctly.
# ═══════════════════════════════════════════════════════════════════

OVERRIDE_KEY_VERSION = 2  # v2: rep name is case/space-insensitive in the key


def _visit_key(customer, visit_date, rep, note) -> str:
    d = str(pd.to_datetime(visit_date, errors="coerce"))[:10]
    rep_norm = " ".join(normalize_arabic(safe_str(rep)).split()).casefold()
    base = "|".join([
        normalize_arabic(safe_str(customer)),
        d,
        rep_norm,
        normalize_arabic(safe_str(note))[:60],
    ])
    return hashlib.md5(base.encode("utf-8")).hexdigest()


def _visit_key_series(df: pd.DataFrame) -> pd.Series:
    """
    Key computation for a classified DataFrame.
    Uses the ORIGINAL customer name when available so that approved name
    merges never break saved manual classifications.
    """
    if "Customer Name Original" in df.columns:
        orig = df["Customer Name Original"].fillna("").astype(str)
        disp = df["Customer Name"].astype(str) if "Customer Name" in df.columns else orig
        cust = np.where(orig != "", orig, disp)
    else:
        cust = df["Customer Name"] if "Customer Name" in df.columns else [""] * len(df)
    date = df["Visit Date"]     if "Visit Date"     in df.columns else [""] * len(df)
    rep  = df["Sales Rep Name"] if "Sales Rep Name" in df.columns else [""] * len(df)
    note = df["Visit Notes"]    if "Visit Notes"    in df.columns else [""] * len(df)
    return pd.Series(
        [_visit_key(c, d, r, n) for c, d, r, n in zip(cust, date, rep, note)],
        index=df.index,
    )


def export_unclassified(classified_df: pd.DataFrame) -> bytes:
    """
    Export Unclassified rows as Excel for manual classification.
    Includes Row Number as the key for matching on re-import.
    """
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    # Get unclassified rows with their original row numbers
    mask = classified_df["Display Status"] == "Unclassified"
    unclass = classified_df[mask].copy()
    unclass.insert(0, "Row Number", unclass.index)

    # Select display columns ("Customer Name Original" keeps the override
    # key stable even after approved name merges)
    show_cols = ["Row Number", "Visit Date", "Customer Name", "Sales Rep Name",
                 "Governorate", "District", "Visit Notes", "Display Status",
                 "Customer Name Original"]
    show_cols = [c for c in show_cols if c in unclass.columns]
    unclass = unclass[show_cols].copy()

    # Add empty Manual Status column
    unclass["Manual Status"] = ""

    # Format dates
    if "Visit Date" in unclass.columns:
        unclass["Visit Date"] = pd.to_datetime(
            unclass["Visit Date"], errors="coerce"
        ).dt.strftime("%Y-%m-%d")

    # Build Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unclassified Visits"
    ws.sheet_view.rightToLeft = True

    # Header style
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    manual_fill = PatternFill("solid", fgColor="E2EFDA")  # Green tint for manual column
    border = Border(
        left=Side(border_style="thin", color="B8CCE4"),
        right=Side(border_style="thin", color="B8CCE4"),
        top=Side(border_style="thin", color="B8CCE4"),
        bottom=Side(border_style="thin", color="B8CCE4"),
    )

    # Write header
    for col_idx, col_name in enumerate(unclass.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font   = header_font
        cell.fill   = manual_fill if col_name == "Manual Status" else header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write data
    for row_idx, row in enumerate(unclass.itertuples(index=False), start=2):
        fill = PatternFill("solid", fgColor="EBF3FB") if row_idx % 2 == 0 else PatternFill()
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            col_name = unclass.columns[col_idx - 1]
            if col_name == "Manual Status":
                cell.fill = PatternFill("solid", fgColor="E2EFDA")
            else:
                cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # Add dropdown validation note
    ws.cell(row=1, column=len(unclass.columns), value="Manual Status").comment = None

    # Auto width
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(50, max(12, max_len + 4))

    # Notes column wider
    ws.column_dimensions["G"].width = 60
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "B2"

    # Add valid statuses note in a separate sheet
    ws2 = wb.create_sheet("القيم المسموح بها")
    ws2.cell(row=1, column=1, value="القيم المسموح بها في عمود Manual Status")
    ws2.cell(row=1, column=1).font = Font(bold=True, color="1F4E79")
    for i, status in enumerate(VALID_STATUSES, start=2):
        ws2.cell(row=i, column=1, value=status)
    ws2.column_dimensions["A"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_overrides(
    override_file,
    classified_df: pd.DataFrame,
) -> tuple[pd.DataFrame, int, list[str]]:
    """
    Import manual classifications from uploaded override file.

    Returns:
        (updated_classified_df, count_changed, errors_list)
    """
    errors = []

    try:
        override_df = pd.read_excel(override_file, engine="openpyxl")
    except Exception as e:
        return classified_df, 0, [f"❌ لا يمكن قراءة الملف: {e}"]

    # Validate required columns
    if "Row Number" not in override_df.columns:
        return classified_df, 0, ["❌ عمود 'Row Number' غير موجود في الملف"]
    if "Manual Status" not in override_df.columns:
        return classified_df, 0, ["❌ عمود 'Manual Status' غير موجود في الملف"]

    # Filter rows with actual manual status
    override_df = override_df[
        override_df["Manual Status"].notna() &
        (override_df["Manual Status"].astype(str).str.strip() != "")
    ].copy()

    if override_df.empty:
        return classified_df, 0, ["⚠️ لا توجد تصنيفات يدوية في الملف — عمود Manual Status فارغ"]

    # Validate status values
    invalid_rows = []
    for _, row in override_df.iterrows():
        status = str(row["Manual Status"]).strip()
        if status not in VALID_STATUSES:
            invalid_rows.append(f"صف {row.get('Row Number', '?')}: قيمة غير معروفة '{status}'")

    if invalid_rows:
        errors.extend(invalid_rows[:10])  # Show max 10 errors
        # Remove invalid rows
        override_df = override_df[
            override_df["Manual Status"].astype(str).str.strip().isin(VALID_STATUSES)
        ]

    # Build stable keys from the override file's own columns
    # (the export always contains Customer Name / Visit Date / Sales Rep Name / Visit Notes)
    key_cols = ["Customer Name", "Visit Date", "Sales Rep Name", "Visit Notes"]
    missing_key_cols = [c for c in key_cols if c not in override_df.columns]
    if missing_key_cols:
        return classified_df, 0, [f"❌ أعمدة ناقصة في الملف: {', '.join(missing_key_cols)}"]

    override_df = override_df.copy()
    override_df["_key"] = _visit_key_series(override_df)

    # Map keys → row positions in classified_df
    updated_df = classified_df.copy()
    data_keys = _visit_key_series(updated_df)
    key_to_idx: dict = {}
    for idx, k in data_keys.items():
        key_to_idx.setdefault(k, []).append(idx)

    count_changed = 0
    applied_rows = []
    for _, row in override_df.iterrows():
        new_status = str(row["Manual Status"]).strip()
        indices = key_to_idx.get(row["_key"], [])
        if not indices:
            errors.append(
                f"⚠️ زيارة غير موجودة في البيانات: {safe_str(row.get('Customer Name'))} — {str(row.get('Visit Date'))[:10]}"
            )
            continue
        for idx in indices:
            updated_df.at[idx, "Display Status"]   = new_status
            updated_df.at[idx, "Suggested Status"] = _display_to_internal(new_status)
            updated_df.at[idx, "Override Source"]  = "Manual"
            updated_df.at[idx, "Confidence Score"] = 100.0
            count_changed += 1
        applied_rows.append({
            "Key":            row["_key"],
            "Customer Name":  safe_str(row.get("Customer Name")),
            "Visit Date":     str(row.get("Visit Date"))[:10],
            "Sales Rep Name": safe_str(row.get("Sales Rep Name")),
            "Manual Status":  new_status,
        })

    # Save overrides to disk for persistence
    if applied_rows:
        _save_overrides(pd.DataFrame(applied_rows))

    return updated_df, count_changed, errors


def _display_to_internal(display: str) -> str:
    """Convert display status to internal key."""
    mapping = {
        "Current Customer":   "current",
        "Potential Customer": "potential",
        "Target Customer":    "target",
        "New Customer":       "new",
        "Former Customer":    "former",
        "Not Interested":     "not_interested",
        "Unclassified":       "unclassified",
    }
    return mapping.get(display, "unclassified")


def _save_overrides(override_df: pd.DataFrame):
    """Save keyed overrides to parquet for persistence."""
    paths = _paths()
    paths["overrides"].parent.mkdir(parents=True, exist_ok=True)

    existing = pd.DataFrame()
    if paths["overrides"].exists():
        try:
            existing = _parquet_to_df(paths["overrides"])
        except Exception:
            existing = pd.DataFrame()

    if not existing.empty and "Key" in existing.columns:
        # Merge: new overrides overwrite existing ones for the same visit
        combined = pd.concat([existing, override_df], ignore_index=True)
        combined = combined.drop_duplicates(subset=["Key"], keep="last")
    else:
        combined = override_df.copy()

    with _write_lock():
        _df_to_parquet(combined, paths["overrides"])
        # Update metadata
        meta = _load_metadata()
        meta["override_count"] = len(combined)
        _save_metadata(meta)


def _rekey_overrides_if_needed(classified_df: pd.DataFrame):
    """
    One-time re-keying when the key formula changes (OVERRIDE_KEY_VERSION).
    The saved classified data already has manual rows marked
    (Override Source == "Manual"), so the overrides file can be rebuilt
    from those rows with fresh keys — nothing is lost.
    """
    try:
        meta = _load_metadata()
        if int(meta.get("override_key_version", 1)) >= OVERRIDE_KEY_VERSION:
            return
        if "Override Source" in classified_df.columns:
            manual = classified_df[classified_df["Override Source"] == "Manual"]
            if not manual.empty:
                keys = _visit_key_series(manual)
                rebuilt = pd.DataFrame({
                    "Key":            keys.values,
                    "Customer Name":  manual["Customer Name"].astype(str).values
                                      if "Customer Name" in manual.columns else "",
                    "Visit Date":     manual["Visit Date"].astype(str).str[:10].values
                                      if "Visit Date" in manual.columns else "",
                    "Sales Rep Name": manual["Sales Rep Name"].astype(str).values
                                      if "Sales Rep Name" in manual.columns else "",
                    "Manual Status":  manual["Display Status"].astype(str).values,
                }).drop_duplicates(subset=["Key"], keep="last")
                with _write_lock():
                    _df_to_parquet(rebuilt, _paths()["overrides"])
        meta["override_key_version"] = OVERRIDE_KEY_VERSION
        _save_metadata(meta)
    except Exception:
        pass  # keep old overrides file untouched on any failure


def _migrate_legacy_overrides(overrides: pd.DataFrame, classified_df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert old row-number overrides to stable visit keys.
    Safe because the saved classified parquet is exactly the dataset
    those row numbers were created against.
    """
    migrated = []
    for _, row in overrides.iterrows():
        try:
            row_num = int(row["Row Number"])
        except (ValueError, TypeError):
            continue
        if row_num not in classified_df.index:
            continue
        src = classified_df.loc[row_num]
        migrated.append({
            "Key": _visit_key(src.get("Customer Name"), src.get("Visit Date"),
                              src.get("Sales Rep Name"), src.get("Visit Notes")),
            "Customer Name":  safe_str(src.get("Customer Name")),
            "Visit Date":     str(src.get("Visit Date"))[:10],
            "Sales Rep Name": safe_str(src.get("Sales Rep Name")),
            "Manual Status":  str(row["Manual Status"]).strip(),
        })
    migrated_df = pd.DataFrame(migrated).drop_duplicates(subset=["Key"], keep="last")
    if not migrated_df.empty:
        with _write_lock():
            _df_to_parquet(migrated_df, _paths()["overrides"])
    return migrated_df


def _apply_saved_overrides(classified_df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Apply any saved overrides to classified_df on load (key-based)."""
    paths = _paths()

    if not paths["overrides"].exists():
        return classified_df, 0

    try:
        overrides = _parquet_to_df(paths["overrides"])
        if overrides.empty:
            return classified_df, 0

        # One-time migration from the legacy row-number format
        if "Key" not in overrides.columns and "Row Number" in overrides.columns:
            overrides = _migrate_legacy_overrides(overrides, classified_df)
            if overrides.empty:
                return classified_df, 0

        updated = classified_df.copy()
        data_keys = _visit_key_series(updated)
        key_to_idx: dict = {}
        for idx, k in data_keys.items():
            key_to_idx.setdefault(k, []).append(idx)

        count = 0
        for _, row in overrides.iterrows():
            new_status = str(row["Manual Status"]).strip()
            if new_status not in VALID_STATUSES:
                continue
            for idx in key_to_idx.get(row.get("Key"), []):
                updated.at[idx, "Display Status"]   = new_status
                updated.at[idx, "Suggested Status"] = _display_to_internal(new_status)
                updated.at[idx, "Override Source"]  = "Manual"
                updated.at[idx, "Confidence Score"] = 100.0
                count += 1

        return updated, count
    except Exception:
        return classified_df, 0


def clear_overrides() -> tuple[bool, str]:
    """Delete all saved overrides."""
    paths = _paths()
    try:
        if paths["overrides"].exists():
            paths["overrides"].unlink()
        meta = _load_metadata()
        meta["override_count"] = 0
        _save_metadata(meta)
        return True, "✅ تم حذف كل التصنيفات اليدوية"
    except Exception as e:
        return False, f"❌ خطأ: {e}"


def clear_all_data() -> tuple[bool, str]:
    """Delete all saved data (reset). Custom keyword rules are kept."""
    try:
        data_dir = get_data_dir()
        for f in data_dir.glob("*.parquet"):
            f.unlink()
        for f in data_dir.glob("*.json"):
            if f.name != "custom_rules.json":
                f.unlink()
        for f in data_dir.glob("*.xlsx"):
            f.unlink()
        return True, "✅ تم حذف كل البيانات المحفوظة (تم الاحتفاظ بالقواعد المخصصة)"
    except Exception as e:
        return False, f"❌ خطأ: {e}"


# ═══════════════════════════════════════════════════════════════════
# CUSTOM KEYWORD RULES (editable from the UI)
# ═══════════════════════════════════════════════════════════════════

def load_custom_rules() -> list[dict]:
    """Load user-defined keyword rules: [{"keyword","status","score"}, ...]"""
    p = _paths()["custom_rules"]
    if not p.exists():
        return []
    try:
        with open(p, "r", encoding="utf-8") as f:
            rules = json.load(f)
        return [r for r in rules
                if isinstance(r, dict) and r.get("keyword") and r.get("status")]
    except Exception:
        return []


def save_custom_rules(rules: list[dict]) -> tuple[bool, str]:
    p = _paths()["custom_rules"]
    try:
        p.parent.mkdir(parents=True, exist_ok=True)
        tmp = p.with_name(p.name + ".tmp")
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(rules, f, ensure_ascii=False, indent=2)
        os.replace(tmp, p)
        return True, "✅ تم حفظ القواعد المخصصة"
    except Exception as e:
        return False, f"❌ خطأ في حفظ القواعد: {e}"


# ═══════════════════════════════════════════════════════════════════
# CUSTOMER NAME MERGES (user-approved unification)
# Names are only merged within the SAME governorate, and only after
# explicit user approval in the review screen.
# ═══════════════════════════════════════════════════════════════════

def load_name_merges() -> pd.DataFrame:
    """Columns: Governorate, Variant (raw name), Canonical (unified name)."""
    p = _paths()["name_merges"]
    if not p.exists():
        return pd.DataFrame(columns=["Governorate", "Variant", "Canonical"])
    try:
        return _parquet_to_df(p)
    except Exception:
        return pd.DataFrame(columns=["Governorate", "Variant", "Canonical"])


def save_name_merges(merges_df: pd.DataFrame) -> tuple[bool, str]:
    try:
        with _write_lock():
            _df_to_parquet(merges_df, _paths()["name_merges"])
        return True, "✅ تم حفظ توحيد الأسماء"
    except Exception as e:
        return False, f"❌ خطأ: {e}"


def apply_name_merges(df: pd.DataFrame) -> pd.DataFrame:
    """
    Apply approved merges in-memory. The original spelling is always kept in
    'Customer Name Original', and is restored first so removing a merge
    takes effect on the next load.
    """
    if df.empty or "Customer Name" not in df.columns:
        return df
    df = df.copy()

    # Restore originals first (makes merge removal reversible)
    if "Customer Name Original" in df.columns:
        restored = df["Customer Name Original"].fillna("").astype(str)
        df["Customer Name"] = np.where(restored != "", restored, df["Customer Name"])
    else:
        df["Customer Name Original"] = df["Customer Name"]

    merges = load_name_merges()
    if merges.empty:
        return df

    gov = df["Governorate"].astype(str) if "Governorate" in df.columns else ""
    mapping = {
        (safe_str(r["Governorate"]), safe_str(r["Variant"])): safe_str(r["Canonical"])
        for _, r in merges.iterrows()
    }
    keys = list(zip(gov, df["Customer Name"].astype(str)))
    df["Customer Name"] = [mapping.get(k, k[1]) for k in keys]
    return df


# ═══════════════════════════════════════════════════════════════════
# PARQUET HELPERS (fast binary format, better than CSV/Excel)
# ═══════════════════════════════════════════════════════════════════

def _df_to_parquet(df: pd.DataFrame, path: Path):
    """Save DataFrame as parquet atomically (write temp file, then swap)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    df_save = df.copy()
    # Convert datetime columns to string to avoid timezone issues
    for col in df_save.select_dtypes(include=["datetime64[ns]", "datetime64[ns, UTC]"]).columns:
        df_save[col] = df_save[col].astype(str)
    tmp = path.with_name(path.name + ".tmp")
    df_save.to_parquet(str(tmp), index=True, engine="pyarrow" if _has_pyarrow() else "fastparquet")
    os.replace(tmp, path)


def _parquet_to_df(path: Path) -> pd.DataFrame:
    """Load DataFrame from parquet."""
    df = pd.read_parquet(str(path), engine="pyarrow" if _has_pyarrow() else "fastparquet")
    # Re-parse Visit Date if it was stored as string
    if "Visit Date" in df.columns and df["Visit Date"].dtype == object:
        df["Visit Date"] = pd.to_datetime(df["Visit Date"], errors="coerce")
    return df


def _has_pyarrow() -> bool:
    try:
        import pyarrow
        return True
    except ImportError:
        return False


# ═══════════════════════════════════════════════════════════════════
# STORAGE STATUS
# ═══════════════════════════════════════════════════════════════════

def storage_status() -> dict:
    """Return storage health info for display in Settings page."""
    data_dir = get_data_dir()
    paths    = _paths()
    meta     = _load_metadata()

    files_exist = {k: v.exists() for k, v in paths.items()}
    total_size  = sum(
        v.stat().st_size for v in paths.values() if v.exists()
    )

    return {
        "data_dir":       str(data_dir),
        "dir_exists":     data_dir.exists(),
        "dir_writable":   _test_write(data_dir),
        "has_data":       files_exist.get("classified", False),
        "has_overrides":  files_exist.get("overrides", False),
        "total_size_kb":  round(total_size / 1024, 1),
        "metadata":       meta,
        "files":          files_exist,
    }


def _test_write(path: Path) -> bool:
    try:
        test = path / ".write_test"
        test.write_text("x")
        test.unlink()
        return True
    except Exception:
        return False
